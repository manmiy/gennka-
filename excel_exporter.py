"""
Excel出力モジュール
原価管理表形式のExcelファイルを生成する
"""

import io
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COLUMNS = [
    {"key": "row_no", "header": "行", "width": 5},
    {"key": "category", "header": "分類", "width": 12},
    {"key": "product_code", "header": "コード", "width": 14},
    {"key": "product_name", "header": "名称", "width": 30},
    {"key": "spec", "header": "仕様", "width": 20},
    {"key": "unit", "header": "単位", "width": 6},
    {"key": "quantity", "header": "数量", "width": 8},
    {"key": "unit_price", "header": "単価", "width": 12},
    {"key": "amount", "header": "金額", "width": 14},
    {"key": "slip_no", "header": "伝票No", "width": 10},
    {"key": "order_no", "header": "注文No", "width": 12},
    {"key": "remarks", "header": "備考", "width": 25},
]

CATEGORY_MAP = {
    "吉野石膏": "ボード類",
    "DAIKEN": "ボード類",
    "大建": "ボード類",
    "TOTO": "住設機器",
    "トクラス": "住設機器",
    "KVK": "給排水設備",
    "LIXIL": "住設機器",
    "リクシル": "住設機器",
    "INAX": "住設機器",
    "サンダイヤ": "設備機器",
    "コロナ": "暖房・給湯",
    "ベストパーツ": "配管部材",
    "JSP": "断熱材",
    "旭ファイバーグラ": "断熱材",
    "旭ファイバー": "断熱材",
    "樋口仕入先": "ケイカル板",
    "城東テクノ": "設備部材",
    "ニチハ": "外壁材",
    "TOTO機": "住設機器",
    "三浦金物": "金物・副資材",
    "日本ライティング": "照明器具",
    "オーデリック": "照明器具",
    "パナソニック": "電気・住設",
    "ルームワン": "カーテン・インテリア",
    "シンコール": "インテリア",
    "タチカワ": "ブラインド・カーテン",
}


def classify_item(maker: str, product_name: str) -> str:
    """メーカー名や品名から適切な分類を推定する"""
    for key, category in CATEGORY_MAP.items():
        if key.lower() in maker.lower():
            return category

    keywords = {
        "ベベルボード": "ボード類",
        "ダイライト": "ボード類",
        "石膏ボード": "ボード類",
        "プラスターボード": "ボード類",
        "ミラフォーム": "断熱材",
        "グラスウール": "断熱材",
        "アクリアウール": "断熱材",
        "断熱": "断熱材",
        "ケイカル板": "ケイカル板",
        "合板": "合板類",
        "コンパネ": "合板類",
        "ベニヤ": "合板類",
        "サッシ": "開口部",
        "窓枠": "開口部",
        "サーモス": "開口部",
        "値引き": "値引き",
        "NEBIKI": "値引き",
        "システムバス": "住設機器",
        "ユニットバス": "住設機器",
        "システムキッチン": "住設機器",
        "キッチン": "住設機器",
        "カップボード": "住設機器",
        "洗面": "住設機器",
        "トイレ": "住設機器",
        "便器": "住設機器",
        "ボイラー": "給湯・暖房",
        "給湯器": "給湯・暖房",
        "オイルタンク": "設備機器",
        "LED": "照明器具",
        "シーリング": "照明器具",
        "ダウンライト": "照明器具",
        "スポットライト": "照明器具",
        "ブラケット": "照明器具",
        "ドレープ": "カーテン",
        "レース": "カーテン",
        "カーテンレール": "カーテン",
        "ロールスクリーン": "インテリア",
        "ブラインド": "インテリア",
        "フサカケ": "カーテン",
        "シリコン": "副資材",
        "コーキング": "副資材",
        "テープ": "副資材",
        "ボルト": "金物",
        "座金": "金物",
        "釘": "金物",
        "ビス": "金物",
        "ステープル": "金物",
        "アンカー": "金物",
        "結束線": "金物",
        "取付施工費": "施工費",
        "施工費": "施工費",
        "諸経費": "諸経費",
        "配送費": "運賃・諸経費",
        "送料": "運賃・諸経費",
    }

    for keyword, category in keywords.items():
        if keyword.lower() in product_name.lower():
            return category

    return ""


def items_to_dataframe(items: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for i, item in enumerate(items, start=1):
        maker = str(item.get("maker", ""))
        product_name = str(item.get("product_name", ""))

        row = {
            "行": i,
            "分類": classify_item(maker, product_name),
            "コード": str(item.get("product_code", "")),
            "名称": product_name,
            "仕様": _build_spec(item),
            "単位": str(item.get("unit", "")),
            "数量": _safe_number(item.get("quantity", "")),
            "単価": _safe_number(item.get("unit_price", "")),
            "金額": _safe_number(item.get("amount", "")),
            "伝票No": str(item.get("slip_no", "")),
            "注文No": str(item.get("order_no", "")),
            "備考": str(item.get("remarks", "")),
            "発注場所": str(item.get("order_location", "")),
        }
        rows.append(row)

    if not rows:
        return pd.DataFrame(
            columns=[
                "行", "分類", "コード", "名称", "仕様", "単位", "数量", "単価", "金額",
                "伝票No", "注文No", "備考", "発注場所",
            ]
        )

    return pd.DataFrame(rows)


def _build_spec(item: Dict[str, Any]) -> str:
    """仕様情報を取得する（明示的なspecがあれば優先し、無ければproduct_code等を使用）"""
    spec = str(item.get("spec", "")).strip()
    if spec:
        return spec
    product_code = str(item.get("product_code", "")).strip()
    return product_code


def _safe_number(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("，", "").strip()
        if cleaned == "" or cleaned == "-":
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def create_excel(
    df: pd.DataFrame,
    title: str = "原価管理表",
    supplier: str = "",
    date_str: str = "",
) -> bytes:
    """
    発注場所ごとに表を分けたExcelファイルを生成する。

    各表は以下の固定列にのみデータを書き込み、それ以外の列は空欄のままにする。
        H列: 名称 / I列: 仕様 / J列: 単位
        W列: 発注数量 / X列: 発注単価 / Y列: 発注金額

    Args:
        df: 明細データのDataFrame（items_to_dataframeの出力。"発注場所"列を含む）
        title: シート先頭に表示するタイトル
        supplier: 仕入先名
        date_str: 請求日

    Returns:
        Excelファイルのバイトデータ
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "発注書"

    # 使用する列（H, I, J, W, X, Y）
    COL_NAME, COL_SPEC, COL_UNIT = 8, 9, 10
    COL_QTY, COL_PRICE, COL_AMOUNT = 23, 24, 25
    FIRST_COL, LAST_COL = COL_NAME, COL_AMOUNT

    header_font = Font(name="Yu Gothic", size=10, bold=True)
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    data_font = Font(name="Yu Gothic", size=10)
    title_font = Font(name="Yu Gothic", size=14, bold=True)
    subtitle_font = Font(name="Yu Gothic", size=10)
    location_font = Font(name="Yu Gothic", size=12, bold=True, color="FFFFFF")
    location_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    subtotal_font = Font(name="Yu Gothic", size=10, bold=True)
    subtotal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 列幅設定
    # A〜G列 (1〜7): 幅1
    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 1

    # H, I, J列 (8, 9, 10): データ列
    ws.column_dimensions[get_column_letter(COL_NAME)].width = 30
    ws.column_dimensions[get_column_letter(COL_SPEC)].width = 20
    ws.column_dimensions[get_column_letter(COL_UNIT)].width = 6

    # K〜V列 (11〜22): 幅1
    for col_idx in range(11, 23):
        ws.column_dimensions[get_column_letter(col_idx)].width = 1

    # W, X, Y列 (23, 24, 25): データ列
    ws.column_dimensions[get_column_letter(COL_QTY)].width = 10
    ws.column_dimensions[get_column_letter(COL_PRICE)].width = 12
    ws.column_dimensions[get_column_letter(COL_AMOUNT)].width = 14

    # Z〜AJ列 (26〜36): 幅1
    for col_idx in range(26, 37):
        ws.column_dimensions[get_column_letter(col_idx)].width = 1

    row = 1

    # タイトル行
    ws.merge_cells(start_row=row, start_column=FIRST_COL, end_row=row, end_column=LAST_COL)
    title_cell = ws.cell(row=row, column=FIRST_COL, value=title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # サブ情報行（仕入先・請求日）
    if supplier or date_str:
        ws.merge_cells(start_row=row, start_column=FIRST_COL, end_row=row, end_column=LAST_COL)
        info_parts = []
        if supplier:
            info_parts.append(f"仕入先: {supplier}")
        if date_str:
            info_parts.append(f"請求日: {date_str}")
        info_cell = ws.cell(row=row, column=FIRST_COL, value="　　".join(info_parts))
        info_cell.font = subtitle_font
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1  # 空行

    # 発注場所ごとにグループ化
    if "発注場所" not in df.columns:
        df = df.copy()
        df["発注場所"] = ""

    location_order: List[str] = []
    location_groups: Dict[str, List[int]] = {}
    unset_indices: List[int] = []

    for idx, loc in zip(df.index, df["発注場所"].fillna("")):
        loc = str(loc).strip()
        if loc == "":
            unset_indices.append(idx)
            continue
        if loc not in location_groups:
            location_groups[loc] = []
            location_order.append(loc)
        location_groups[loc].append(idx)

    if unset_indices:
        location_groups["（発注場所未設定）"] = unset_indices
        location_order.append("（発注場所未設定）")

    if not location_order:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    for location in location_order:
        indices = location_groups[location]
        group_df = df.loc[indices]

        # 発注場所の見出し行
        ws.merge_cells(start_row=row, start_column=FIRST_COL, end_row=row, end_column=LAST_COL)
        loc_cell = ws.cell(row=row, column=FIRST_COL, value=f"発注場所: {location}")
        loc_cell.font = location_font
        loc_cell.fill = location_fill
        loc_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        # ヘッダー行
        header_map = {
            COL_NAME: "名称",
            COL_SPEC: "仕様",
            COL_UNIT: "単位",
            COL_QTY: "発注数量",
            COL_PRICE: "発注単価",
            COL_AMOUNT: "発注金額",
        }
        for col_idx, header in header_map.items():
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22
        row += 1

        # データ行
        data_start_row = row
        for _, item_row in group_df.iterrows():
            values = {
                COL_NAME: item_row.get("名称", ""),
                COL_SPEC: item_row.get("仕様", ""),
                COL_UNIT: item_row.get("単位", ""),
                COL_QTY: item_row.get("数量", ""),
                COL_PRICE: item_row.get("単価", ""),
                COL_AMOUNT: item_row.get("金額", ""),
            }
            for col_idx, value in values.items():
                if pd.isna(value):
                    value = ""
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in (COL_QTY, COL_PRICE, COL_AMOUNT):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, (int, float)) and value != "":
                        cell.number_format = "#,##0"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 20
            row += 1

        data_end_row = row - 1

        # 小計行
        if data_end_row >= data_start_row:
            subtotal_label_cell = ws.cell(row=row, column=COL_NAME, value="小計")
            subtotal_label_cell.font = subtotal_font
            subtotal_label_cell.fill = subtotal_fill
            subtotal_label_cell.border = thin_border
            subtotal_label_cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx in (COL_SPEC, COL_UNIT):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = subtotal_fill
                cell.border = thin_border

            qty_letter = get_column_letter(COL_QTY)
            amt_letter = get_column_letter(COL_AMOUNT)

            qty_cell = ws.cell(
                row=row, column=COL_QTY,
                value=f"=SUM({qty_letter}{data_start_row}:{qty_letter}{data_end_row})",
            )
            qty_cell.font = subtotal_font
            qty_cell.fill = subtotal_fill
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(horizontal="right", vertical="center")
            qty_cell.number_format = "#,##0"

            price_cell = ws.cell(row=row, column=COL_PRICE)
            price_cell.fill = subtotal_fill
            price_cell.border = thin_border

            amt_cell = ws.cell(
                row=row, column=COL_AMOUNT,
                value=f"=SUM({amt_letter}{data_start_row}:{amt_letter}{data_end_row})",
            )
            amt_cell.font = subtotal_font
            amt_cell.fill = subtotal_fill
            amt_cell.border = thin_border
            amt_cell.alignment = Alignment(horizontal="right", vertical="center")
            amt_cell.number_format = "#,##0"

            ws.row_dimensions[row].height = 22
            row += 1

        row += 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
